import asyncio

from fastapi import BackgroundTasks, Form
from fastapi import FastAPI
from fastapi import WebSocket
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook
from sh import tail
from sse_starlette.sse import EventSourceResponse
from starlette.middleware.cors import CORSMiddleware
from starlette.requests import Request
from starlette.responses import FileResponse, RedirectResponse

from app_moules import *

from database import db_connect, LOGFILE

app = FastAPI()
templates = Jinja2Templates(directory="templates")
connect = db_connect()
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

d4 = {}

sql_2 = 'SELECT COUNT(id) FROM states'
cd = connect.query_database_df(sql_2)
print(cd['COUNT(id)'][0])

sql_1 = 'SELECT * FROM states'
c_1 = connect.query_database_df(sql_1)
for k in range(0, len(c_1)):
    d4[c_1['id'][k]] = c_1['state_name'][k]


@app.get('/')
def enter(request: Request):
    return templates.TemplateResponse("base.html", {"request": request})


@app.get("/scrape_data")
async def home(background_tasks: BackgroundTasks):
    background_tasks.add_task(collect_data)
    #response = RedirectResponse(url='/stream-logs')
    response= {"message":"process running"}
    return response


@app.get("/update_scrape_data")
async def update(background_tasks: BackgroundTasks):
    #response = RedirectResponse(url='/stream-logs')
    response = {"message": "process running"}
    return response


@app.get('/enter_data')
def enter(request: Request, dt: dict = None):
    if dt is None:
        dt = d4
    return templates.TemplateResponse("form.html", {"request": request, "states": dt})


@app.post('/submit_data')
async def download(state: str = Form(...)):
    print(state)
    d1, d2 = return_states()
    wb = Workbook()
    total_ngos = d1[state]['total_ngos']
    for key, value in d1[state]['pages'].items():
        if key == "1":
            ws1 = wb.active
            ws1.title = "Page_no" + str(key)
        else:
            ws1 = wb.create_sheet("Page_no" + str(key))
        ws1.cell(row=1, column=1, value='row')
        ws1.cell(row=1, column=2, value="name")
        ws1.cell(row=1, column=3, value="reg_date")
        ws1.cell(row=1, column=4, value="address")
        ws1.cell(row=1, column=5, value="phone no")
        ws1.cell(row=1, column=6, value="email")
        i = 2
        for k in value:
            ws1.cell(row=i, column=1, value=k['ngo_row'])
            ws1.cell(row=i, column=2, value=k['ngo_name'])
            ws1.cell(row=i, column=3, value=k['reg_date'])
            ws1.cell(row=i, column=4, value=k['ngo_address'])
            ws1.cell(row=i, column=5, value=k['ngo_mobile'])
            ws1.cell(row=i, column=6, value=k['ngo_email'])
            i = i + 1
    wb.save("./ngos.xlsx")
    return FileResponse("./ngos.xlsx")


@app.get('/enter_data_2')
def enter_2(request: Request, dt: dict = None, dw: dict = None):
    if dt is None:
        dt = d4
    s7 = {}
    sql = 'SELECT * FROM states'
    c = connect.query_database_df(sql)
    for r in range(0, len(c)):
        s7[c['state_name'][r]] = c['pages'][r]
    if dw is None:
        dw = s7
    return templates.TemplateResponse("form_2.html", {"request": request, "states": dt, 'list': dw})


@app.post('/submit_data_2')
async def download_2(request: Request, state: str = Form(...), page: str = Form(...), de: dict = None):
    print(state)
    print(page)
    d1, d2 = return_states()
    if de is None:
        de = d1[state]['pages'][page]
    print(de)
    return templates.TemplateResponse("result.html", {"request": request, "dt": de,"state":state,"page":page})


async def logGenerator(request):
    for line in tail("-f", LOGFILE, _iter=True):
        if await request.is_disconnected():
            print("client disconnected!!!")
            break
        yield line


@app.get('/stream-logs')
async def run(request: Request):
    event_generator = logGenerator(request)
    return EventSourceResponse(event_generator)


@app.websocket("/see_status")
async def status(websocket: WebSocket):
    d1, d2 = return_states()
    await websocket.accept()
    while True:
        await asyncio.sleep(0.1)
        # for i in d2:
        payload = d2
        # payload = requests.get("http://localhost:8000/get_status_data")
        await websocket.send_json(payload)


@app.get("/get_status_data")
async def get_status_data(request: Request, d21: dict = None, dt: dict = None):
    d1, d2 = return_states()
    if d21 is None:
        d21 = d2
    if dt is None:
        dt = d4
    return templates.TemplateResponse("status.html", {"request": request, "dt": d21, "dr": dt})


@app.post('/submit_data_3')
async def download_3(request: Request, state: str = Form(...), status: str = Form(...), de: dict = None):
    d1, d2 = return_states()
    print(state)
    print(status)
    if de is None:
        de = d2[state][status]
    print(de)

    return templates.TemplateResponse("part_status.html", {"request": request, "dt": de,"state":state,"status":status})
