import os
import time

from bs4 import BeautifulSoup as soup
from flask import Flask, render_template, request
from openpyxl import Workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

app = Flask(__name__)


class NGO:
    def collect(self, url, options, wb, sh1, filename, start, rows):
        l1 = []
        d1 = {}
        # driver = webdriver.Chrome(executable_path="chromedriver.exe", options=options)
        driver = webdriver.Chrome(executable_path=os.environ.get("CHROMEDRIVER_PATH"), chrome_options=options)
        for i in range(int(start), int(rows) + 1):
            # print(str(i))
            driver.get(url)
            # driver.implicitly_wait(10)
            all = driver.find_element_by_xpath("/html/body/div[9]/div[1]/div[3]/div/div/div[2]/table/tbody/tr[" + str(i) + "]/td[2]/a")
            all.click()
            time.sleep(10)
            overlay = WebDriverWait(driver, 10).until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "div.blockUI.blockOverlay")))
            if overlay:
                close = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, "//*[@id='ngo_info_modal']/div[2]/div/div[1]/button/span")))
                close.click()
                time.sleep(2)
            ngo = driver.find_element_by_id('ngo_info_modal')
            result = ngo.get_attribute('innerHTML')
            sp = soup(result, 'html.parser')
            name = sp.find(id="ngo_name_title")
            add = sp.find(id="address")
            date = sp.find(id="ngo_reg_date")
            mobile = sp.find(id="mobile_n")
            email = sp.find(id="email_n")
            line = str(i + 1), name.text, add.text, date.text, mobile.text, email.text
            # print(line)
            sh1.cell(row=i + 1, column=1, value=name.text)
            sh1.cell(row=i + 1, column=2, value=date.text)
            sh1.cell(row=i + 1, column=3, value=add.text)
            sh1.cell(row=i + 1, column=4, value=mobile.text)
            sh1.cell(row=i + 1, column=5, value=email.text)
            wb.save(filename + ".xls")
            print(line)
            d1["id"] = str(i)
            d1["name "] = name.text
            d1["regdate"] = date.text
            d1["address"] = add.text
            d1["mobile"] = mobile.text
            d1["email"] = email.text
            l1.insert(i, d1.copy)
        print("\n")
        print("---------------START COPYING------------")
        print("\n")
        print(l1)
        print("\n")
        print("---------------END COPYING------------")
        driver.close()


@app.route('/')
def test():
    return render_template("form.html")


@app.route('/scrape', methods=["GET", "POST"])
def scrape():
    # user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.131 Safari/537.36"
    options = webdriver.ChromeOptions()
    # options.add_argument(f'user-agent={user_agent}')
    options.binary_location = os.environ.get("GOOGLE_CHROME_BIN")
    options.headless = True
    options.add_argument("--window-size=1920,1080")
    options.add_argument('--ignore-certificate-errors')
    options.add_argument('--allow-running-insecure-content')
    options.add_argument("--disable-extensions")
    options.add_argument("--proxy-server='direct://'")
    options.add_argument("--proxy-bypass-list=*")
    options.add_argument("--start-maximized")
    options.add_argument('--disable-gpu')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--no-sandbox')
    wb = Workbook()
    sh1 = wb.active
    sh1.cell(row=1, column=1, value="name")
    sh1.cell(row=1, column=2, value="regdate")
    sh1.cell(row=1, column=3, value="address")
    sh1.cell(row=1, column=4, value="phone no")
    sh1.cell(row=1, column=5, value="email")
    url = request.form.get("url")
    start = request.form.get("start")
    rows = request.form.get("Number")
    filename = request.form.get("page")
    wb.save(filename + ".xls")
    if request.method == "POST":
        ngo.collect(url, options, wb, sh1, filename, start, rows)
        response = "success"
        return response


ngo = NGO()
if __name__ == "__main__":
    app.run(debug=True)
